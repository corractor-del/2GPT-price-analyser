# -*- coding: utf-8 -*-
"""
Avito Price Analyzer — версия с обходом защиты:
- Повторы запросов с экспоненциальной паузой и сменой хоста (www ↔ m).
- Поддержка cookies.txt (формат Netscape) — можно загрузить куки из своего браузера для avito.ru.
- Перетаскивание Excel-файла (drag & drop) + кнопка "Стоп" (отмена) и уникальные имена выходных файлов.
- Прогресс, статусы "Готово/Отменено".

Формат Excel без шапки:
A — бренд, B — модель/характеристики, C — закупочная цена ₽.
Выход:
D — средняя цена ₽, E — наценка % ((D-C)/C*100), F — ссылка на минимальное.
"""
import os, sys, re, time, math, random, threading, queue, webbrowser
from dataclasses import dataclass
from typing import List, Optional, Tuple

import requests
from bs4 import BeautifulSoup

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# GUI + drag&drop
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    TKDND = True
except Exception:
    TKDND = False

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from http.cookiejar import MozillaCookieJar

APP_TITLE = "Avito Price Analyzer"
RESULT_SUFFIX = "_analyzed"
REQUEST_TIMEOUT = 20
REQUEST_DELAY_RANGE = (0.8, 1.6)
RETRY_ATTEMPTS = 4  # попыток на каждый поиск
SCRAPE_LIMIT = 120  # карточек со страницы поиска (сырьё)
FILTER_TAKE = 40    # сколько учитываем в среднем

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
]

STOPWORDS_RU = {"и","или","а","для","с","на","в","из","к","по","без","до","от","что",
    "цвет","новый","б/у","бу","есть","нет","про","про-","встроенный","версия"}

@dataclass
class Listing:
    title: str
    url: str
    price_rub: Optional[int]

def normalize_text(s: str) -> str:
    s = s.lower().replace("чёр", "чер")
    s = re.sub(r"(\d+)\s*(гб|gb)", lambda m: f"{m.group(1)}gb", s)
    s = re.sub(r"(\d+)\s*(тб|tb)", lambda m: f"{m.group(1)}tb", s)
    s = re.sub(r"[^a-z0-9а-яё\s\-]+", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def extract_tokens(brand: str, name: str) -> List[str]:
    full = normalize_text(f"{brand} {name}")
    raw = re.findall(r"[a-z0-9а-яё\-]+", full, flags=re.IGNORECASE)
    toks, seen = [], set()
    for t in raw:
        if len(t) <= 1 or t in STOPWORDS_RU: continue
        if t not in seen: seen.add(t); toks.append(t)
    return toks

def build_search_urls(query: str) -> List[str]:
    from urllib.parse import quote_plus
    q = quote_plus(query)
    return [f"https://www.avito.ru/rossiya?q={q}", f"https://m.avito.ru/rossiya?q={q}"]

def fetch_html(url: str, session: requests.Session, log=None) -> Optional[str]:
    headers = {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.6",
        "Connection": "keep-alive",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "DNT": "1",
        "Upgrade-Insecure-Requests": "1",
    }
    try:
        r = session.get(url, headers=headers, timeout=REQUEST_TIMEOUT)
        if r.status_code == 200 and "<html" in r.text.lower():
            return r.text
    except requests.RequestException as e:
        if log: log(f"      сеть: {e}")
    return None

def parse_cards_from_search(html: str) -> List[Listing]:
    soup = BeautifulSoup(html, "lxml")
    txt = soup.get_text(" ", strip=True).lower()
    if "не робот" in txt or "captcha" in txt or "подтвердите" in txt:
        return []
    cands = []
    cands.extend(soup.select('div[data-marker="item"]'))
    cands.extend(soup.select('div[class*="iva-item"]'))
    cands.extend(soup.select('article'))
    uniq, seen = [], set()
    for c in cands:
        key = c.get("data-item-id") or str(hash(c.get_text(" ", strip=True)))[:16]
        if key not in seen:
            seen.add(key); uniq.append(c)
    res = []
    for n in uniq:
        a = n.select_one('a[data-marker="item-title"]') or n.select_one('a[class*="link-link"]') or n.find('a', href=True)
        if not a: continue
        title = a.get_text(" ", strip=True) or None
        href = a.get("href")
        if not title or not href: continue
        if not href.startswith("http"):
            if href.startswith("/"):
                href = "https://www.avito.ru" + href
            else:
                href = "https://www.avito.ru/" + href
        price = None
        mp = n.select_one('meta[itemprop="price"]')
        if mp and mp.get("content") and mp["content"].isdigit():
            price = int(mp["content"])
        else:
            p = n.select_one('[data-marker="item-price"]') or n.select_one('span[itemprop="price"]') or n.select_one('strong[class*="price"]')
            if p:
                digs = re.findall(r"\d+", p.get_text(" ", strip=True))
                if digs: price = int("".join(digs))
        res.append(Listing(title=title, url=href, price_rub=price))
    return res[:SCRAPE_LIMIT]

def score_title(title: str, tokens: List[str]) -> float:
    if not tokens: return 0.0
    t = normalize_text(title); hits = sum(1 for x in tokens if x in t)
    return hits / len(tokens)

def read_cookies(path: str):
    cj = MozillaCookieJar(); cj.load(path, ignore_discard=True, ignore_expires=True); return cj

def unique_output_path(base_no_ext: str, ext: str) -> str:
    path = f"{base_no_ext}{ext}"; n = 1
    while os.path.exists(path):
        path = f"{base_no_ext} ({n}){ext}"; n += 1
    return path

def make_output_path(input_path: str) -> str:
    base, _ = os.path.splitext(input_path)
    return unique_output_path(base + RESULT_SUFFIX, ".xlsx")

def apply_row_fill(ws, row_idx: int, fill: PatternFill, max_col: int):
    for col in range(1, max_col + 1):
        ws.cell(row=row_idx, column=col).fill = fill

def save_with_formatting(df: pd.DataFrame, out_path: str):
    df.to_excel(out_path, header=False, index=False)
    wb = load_workbook(out_path); ws = wb.active
    for i, w in enumerate([22,54,14,16,12,60], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(1, ws.max_row+1):
        ws.cell(row=r, column=3).number_format = '#,##0" ₽"'
        ws.cell(row=r, column=4).number_format = '#,##0" ₽"'
        ws.cell(row=r, column=5).number_format = '0.00" %"'
        ev = ws.cell(row=r, column=5).value
        if isinstance(ev,(int,float)):
            if 5.0 <= ev < 10.0:
                fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"); apply_row_fill(ws, r, fill, 6)
            elif ev >= 10.0:
                fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"); apply_row_fill(ws, r, fill, 6)
    wb.save(out_path)

class CancelToken:
    def __init__(self): self._flag=False
    def cancel(self): self._flag=True
    def is_cancelled(self): return self._flag

def process_excel(input_path: str, log, progress, status_var: tk.StringVar, cancel: 'CancelToken', cookies_path: Optional[str]) -> Optional[str]:
    try:
        df = pd.read_excel(input_path, header=None)
    except Exception as e:
        log(f"Ошибка чтения файла: {e}"); return None
    if df.shape[1] < 3:
        log("Ошибка: минимум 3 колонки (A,B,C)."); return None

    def to_float(x):
        try:
            if pd.isna(x): return math.nan
            s = str(x).replace(" ", "").replace("\xa0","").replace(",",".")
            s = re.sub(r"[^0-9.\-]","",s); return float(s) if s else math.nan
        except: return math.nan

    session = requests.Session()
    if cookies_path and os.path.exists(cookies_path):
        try:
            cj = read_cookies(cookies_path); session.cookies = cj; log("✓ Куки загружены.")
        except Exception as e:
            log(f"⚠️ Не удалось загрузить cookies: {e}")

    total_rows = len(df)
    for idx, row in df.iterrows():
        if cancel.is_cancelled():
            status_var.set("Отменено"); progress["value"]=0; progress.update_idletasks(); return None

        brand = str(row.iloc[0]) if not pd.isna(row.iloc[0]) else ""
        name  = str(row.iloc[1]) if not pd.isna(row.iloc[1]) else ""
        cost_c = to_float(row.iloc[2])

        progress["value"] = int((idx / max(1,total_rows))*100); progress.update_idletasks()
        tokens = extract_tokens(brand, name)
        query = " ".join(tokens[:8]) if tokens else f"{brand} {name}".strip()
        log(f"[{idx+1}] Ищу: {query}")

        # надёжный фетч с ретраями и сменой хоста
        search_html = None
        urls = build_search_urls(query)
        for attempt in range(1, RETRY_ATTEMPTS+1):
            for u in urls:
                if cancel.is_cancelled(): status_var.set("Отменено"); progress["value"]=0; return None
                log(f"   попытка {attempt}/{RETRY_ATTEMPTS}: {u}")
                search_html = fetch_html(u, session, log=log)
                if search_html: break
                time.sleep((2 ** (attempt-1)) + random.random())
            if search_html: break

        if not search_html:
            log("   ⚠️ Не удалось получить результаты (защита/сеть). Открою страницу в браузере — пройдите капчу, затем нажмите 'Старт' снова.")
            try: webbrowser.open(urls[0])
            except: pass
            df.loc[idx,3]=None; df.loc[idx,4]=None; df.loc[idx,5]=None
            continue

        cards = parse_cards_from_search(search_html)
        log(f"   Сырых карточек: {len(cards)}")
        if not cards:
            log("   ⚠️ Похоже на капчу или пустую выдачу."); df.loc[idx,3]=None; df.loc[idx,4]=None; df.loc[idx,5]=None; continue

        cards.sort(key=lambda li: score_title(li.title, tokens), reverse=True)
        take = [li for li in cards[:FILTER_TAKE] if li.price_rub]
        if not take:
            log("   ⚠️ Есть карточки, но без цен на листинге."); df.loc[idx,3]=None; df.loc[idx,4]=None; df.loc[idx,5]=None; continue

        prices = [li.price_rub for li in take]
        avg_price = round(sum(prices)/len(prices),2)
        cheapest = min(take, key=lambda li: li.price_rub)
        df.loc[idx, 3] = avg_price
        df.loc[idx, 4] = ((avg_price - cost_c) / cost_c * 100.0) if (avg_price and cost_c and not math.isnan(cost_c)) else None
        df.loc[idx, 5] = cheapest.url if cheapest and cheapest.url else None
        log(f"   ✓ Учтено: {len(prices)}; средняя: {avg_price} ₽; мин: {cheapest.price_rub} ₽")

        time.sleep(random.uniform(*REQUEST_DELAY_RANGE))

    out_path = make_output_path(input_path)
    save_with_formatting(df, out_path)
    progress["value"]=100; progress.update_idletasks(); status_var.set("Готово")
    return out_path

# ---------------- GUI ----------------

class AppBase(tk.Tk if not TKDND else __import__('tkinterdnd2').TkinterDnD.Tk): pass

class App(AppBase):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.iconphoto(False, tk.PhotoImage(width=1, height=1))  # иконка в exe через PyInstaller
        self.geometry("900x700")

        self.input_path = None
        self.cookies_path = None
        self.output_path = None
        self.cancel_token = CancelToken()
        self.task_queue = queue.Queue()
        self.status_var = tk.StringVar(value="—")

        self.create_widgets()
        self.after(100, self.process_queue)

    def create_widgets(self):
        pad = 10

        top = ttk.Frame(self, padding=pad); top.pack(fill="x")
        ttk.Label(top, text="Файл Excel (A=бренд, B=модель/характеристики, C=закупка ₽):").pack(anchor="w")
        row = ttk.Frame(top); row.pack(fill="x", pady=(6,0))
        self.path_var = tk.StringVar()
        self.path_entry = ttk.Entry(row, textvariable=self.path_var); self.path_entry.pack(side="left", fill="x", expand=True)
        ttk.Button(row, text="Выбрать файл…", command=self.choose_file).pack(side="left", padx=(6,0))
        if TKDND:
            from tkinterdnd2 import DND_FILES
            self.path_entry.drop_target_register(DND_FILES)
            self.path_entry.dnd_bind("<<Drop>>", self.on_drop)

        cookie_row = ttk.Frame(self, padding=pad); cookie_row.pack(fill="x")
        self.cookie_var = tk.StringVar()
        ttk.Label(cookie_row, text="Cookies (cookies.txt, опционально):").pack(anchor="w")
        c_r = ttk.Frame(cookie_row); c_r.pack(fill="x", pady=(6,0))
        ttk.Entry(c_r, textvariable=self.cookie_var).pack(side="left", fill="x", expand=True)
        ttk.Button(c_r, text="Загрузить cookies…", command=self.choose_cookies).pack(side="left", padx=(6,0))

        btns = ttk.Frame(self, padding=pad); btns.pack(fill="x")
        self.btn_start = ttk.Button(btns, text="Старт", command=self.on_start, state="disabled"); self.btn_start.pack(side="left")
        self.btn_stop  = ttk.Button(btns, text="Стоп",  command=self.on_stop); self.btn_stop.pack(side="left", padx=(6,0))
        ttk.Button(btns, text="Закрыть", command=self.on_close).pack(side="right")

        prog = ttk.Frame(self, padding=pad); prog.pack(fill="x")
        ttk.Label(prog, text="Ход обработки:").pack(anchor="w")
        self.progress = ttk.Progressbar(prog, orient="horizontal", mode="determinate"); self.progress.pack(fill="x", pady=(6,0))
        ttk.Label(prog, textvariable=self.status_var).pack(anchor="e")

        logf = ttk.Frame(self, padding=pad); logf.pack(fill="both", expand=True)
        ttk.Label(logf, text="Лог:").pack(anchor="w")
        self.log = tk.Text(logf, height=18, wrap="word"); self.log.pack(fill="both", expand=True)

        out = ttk.Frame(self, padding=pad); out.pack(fill="x")
        ttk.Label(out, text="Итоговый файл:").pack(anchor="w")
        self.out_var = tk.StringVar(value="—")
        o_r = ttk.Frame(out); o_r.pack(fill="x", pady=(6,0))
        ttk.Entry(o_r, textvariable=self.out_var).pack(side="left", fill="x", expand=True)
        ttk.Button(o_r, text="Открыть папку", command=self.open_output_folder).pack(side="left", padx=(6,0))

    def log(self, text: str):
        self.log.insert(tk.END, text + "\n"); self.log.see(tk.END); self.log.update_idletasks()

    def on_drop(self, event):
        path = event.data.strip("{}")
        if os.path.isfile(path):
            self.input_path = path; self.path_var.set(path); self.btn_start["state"]="normal"

    def choose_file(self):
        path = filedialog.askopenfilename(title="Выберите Excel-файл", filetypes=[("Excel files","*.xlsx *.xls")])
        if path:
            self.input_path = path; self.path_var.set(path); self.btn_start["state"]="normal"

    def choose_cookies(self):
        path = filedialog.askopenfilename(title="Выберите cookies.txt", filetypes=[("Cookies","*.txt"),("Все файлы","*.*")])
        if path:
            self.cookies_path = path; self.cookie_var.set(path)

    def on_start(self):
        if not self.input_path:
            messagebox.showwarning("Внимание","Сначала выберите Excel-файл."); return
        self.progress["value"]=0; self.out_var.set("—"); self.output_path=None; self.log.delete("1.0", tk.END)
        self.status_var.set("В работе…")
        self.btn_start["state"]="disabled"; self.cancel_token = CancelToken()

        def worker():
            try:
                out = process_excel(self.input_path, self.log, self.progress, self.status_var, self.cancel_token, self.cookies_path)
                def ok():
                    if out:
                        self.output_path = out; self.out_var.set(out)
                        if self.status_var.get() != "Отменено": self.status_var.set("Готово")
                        messagebox.showinfo("Статус", self.status_var.get())
                    else:
                        if self.status_var.get() == "Отменено":
                            messagebox.showinfo("Статус", "Отменено")
                        else:
                            messagebox.showwarning("Готово","Завершено с предупреждениями. См. лог.")
                self.task_queue.put(ok)
            except Exception as e:
                def err(): messagebox.showerror("Ошибка", f"Непредвиденная ошибка: {e}")
                self.task_queue.put(err)
            finally:
                def fin(): self.btn_start["state"]="normal"
                self.task_queue.put(fin)

        threading.Thread(target=worker, daemon=True).start()

    def on_stop(self): self.cancel_token.cancel()

    def process_queue(self):
        try:
            while True: self.task_queue.get_nowait()()
        except queue.Empty: pass
        self.after(100, self.process_queue)

    def open_output_folder(self):
        if not self.output_path or not os.path.exists(self.output_path):
            messagebox.showinfo("Инфо","Итоговый файл пока не создан."); return
        folder = os.path.dirname(self.output_path) or "."
        if sys.platform.startswith("win"): os.startfile(folder)
        elif sys.platform == "darwin": os.system(f'open "{folder}"')
        else: os.system(f'xdg-open "{folder}"')

    def on_close(self): self.destroy()

def main(): App().mainloop()
if __name__ == "__main__": main()
